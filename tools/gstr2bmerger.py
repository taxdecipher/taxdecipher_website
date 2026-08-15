from flask import (
    Blueprint,
    render_template,
    request,
    jsonify,
    send_file
)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from werkzeug.utils import secure_filename

import os
import tempfile
import uuid
import threading
import shutil
import re
from datetime import datetime


# ============================================================
# BLUEPRINT
# ============================================================

gstr2bmerger_bp = Blueprint(
    "gstr2bmerger",
    __name__
)


# ============================================================
# PATHS
# ============================================================

BASE_DIR = os.path.dirname(
    os.path.dirname(
        os.path.abspath(__file__)
    )
)

OUTPUT_FILE = os.path.join(
    BASE_DIR,
    "gstr2bmerged_taxdecipher.xlsx"
)


# ============================================================
# TASK STORAGE
# ============================================================

TASKS = {}

TASK_LOCK = threading.Lock()


# ============================================================
# REQUIRED SHEETS
# ============================================================

REQUIRED_SHEETS = [
    "Read me",
    "B2B",
    "B2BA",
    "B2B-CDNR",
    "B2B-CDNRA"
]


# ============================================================
# OUTPUT HEADERS
# ============================================================

B2B_HEADERS = [
    "Month",
    "GSTIN of supplier",
    "Trade/Legal name",
    "Invoice number",
    "Invoice type",
    "Invoice Date",
    "Invoice Value(₹)",
    "Place of supply",
    "Supply Attract Reverse Charge",
    "Taxable Value",
    "IGST",
    "CGST",
    "SGST",
    "Cess",
    "Total",
    "GSTR-1/IFF/GSTR-5 Period",
    "GSTR-1/IFF/GSTR-5 Filing Date",
    "ITC Availability",
    "Reason",
    "Applicable % of Tax Rate",
    "Source",
    "IRN",
    "IRN Date"
]


B2BA_HEADERS = [
    "Month",
    "Invoice number",
    "Invoice Date",
    "GSTIN of supplier",
    "Trade/Legal name",
    "Invoice number",
    "Invoice type",
    "Invoice Date",
    "Invoice Value(₹)",
    "Place of supply",
    "Supply Attract Reverse Charge",
    "Taxable Value (₹)",
    "Integrated Tax(₹)",
    "Central Tax(₹)",
    "State/UT Tax(₹)",
    "Cess(₹)",
    "Total",
    "Whether ITC to be reduced (Taxpayer's Input)",
    "Integrated Tax(₹)",
    "Central Tax(₹)",
    "State/UT Tax(₹)",
    "Cess(₹)",
    "Remarks",
    "GSTR-1/IFF/GSTR-5 Period",
    "GSTR-1/IFF/GSTR-5 Filing Date",
    "ITC Availability",
    "Reason",
    "Applicable % of Tax Rate"
]


B2B_CDNR_HEADERS = [
    "Month",
    "GSTIN of supplier",
    "Trade/Legal name",
    "Note number",
    "Note type",
    "Note Supply type",
    "Note date",
    "Note Value (₹)",
    "Place of supply",
    "Supply Attract Reverse Charge",
    "Taxable Value (₹)",
    "Integrated Tax(₹)",
    "Central Tax(₹)",
    "State/UT Tax(₹)",
    "Cess(₹)",
    "Total",
    "Whether ITC to be reduced (Taxpayer's Input)",
    "Integrated Tax(₹)",
    "Central Tax(₹)",
    "State/UT Tax(₹)",
    "Cess(₹)",
    "Remarks",
    "GSTR-1/IFF/GSTR-5 Period",
    "GSTR-1/IFF/GSTR-5 Filing Date",
    "ITC Availability",
    "Reason",
    "Applicable % of Tax Rate",
    "Source",
    "IRN",
    "IRN Date"
]


B2B_CDNRA_HEADERS = [
    "Month",
    "Note type",
    "Note number",
    "Note date",
    "GSTIN of supplier",
    "Trade/Legal name",
    "Note number",
    "Note type",
    "Note Supply type",
    "Note date",
    "Note Value (₹)",
    "Place of supply",
    "Supply Attract Reverse Charge",
    "Taxable Value (₹)",
    "Integrated Tax(₹)",
    "Central Tax(₹)",
    "State/UT Tax(₹)",
    "Cess(₹)",
    "Total",
    "Whether ITC to be reduced (Taxpayer's Input)",
    "Integrated Tax(₹)",
    "Central Tax(₹)",
    "State/UT Tax(₹)",
    "Cess(₹)",
    "Remarks",
    "GSTR-1/IFF/GSTR-5 Period",
    "GSTR-1/IFF/GSTR-5 Filing Date",
    "ITC Availability",
    "Reason",
    "Applicable % of Tax Rate"
]


# ============================================================
# NORMALIZATION
# ============================================================

def normalize_text(value):

    if value is None:
        return ""

    return (
        str(value)
        .replace("\u00a0", " ")
        .strip()
        .replace("\n", " ")
    )


# ============================================================
# NUMBER CONVERSION
# ============================================================

def number(value):

    if value is None:
        return 0

    if isinstance(value, (int, float)):
        return value

    try:

        text = str(value).strip()

        if text == "":
            return 0

        text = text.replace(",", "")

        return float(text)

    except Exception:

        return 0


# ============================================================
# DATE CONVERSION
# ============================================================

def parse_month(month_name):

    month_name = normalize_text(month_name)

    months = {
        "january": 1,
        "jan": 1,

        "february": 2,
        "feb": 2,

        "march": 3,
        "mar": 3,

        "april": 4,
        "apr": 4,

        "may": 5,

        "june": 6,
        "jun": 6,

        "july": 7,
        "jul": 7,

        "august": 8,
        "aug": 8,

        "september": 9,
        "sep": 9,
        "sept": 9,

        "october": 10,
        "oct": 10,

        "november": 11,
        "nov": 11,

        "december": 12,
        "dec": 12
    }

    return months.get(
        month_name.lower()
    )


def get_period_from_readme(workbook):

    ws = workbook["Read me"]

    financial_year = normalize_text(
        ws["C4"].value
    )

    month_name = normalize_text(
        ws["C5"].value
    )

    if not financial_year:
        raise ValueError(
            "Financial Year not found in Read me!C4"
        )

    if not month_name:
        raise ValueError(
            "Tax Period not found in Read me!C5"
        )

    match = re.search(
        r"(\d{4})\s*-\s*(\d{2,4})",
        financial_year
    )

    if not match:
        raise ValueError(
            f"Invalid financial year: {financial_year}"
        )

    start_year = int(
        match.group(1)
    )

    month_number = parse_month(
        month_name
    )

    if not month_number:
        raise ValueError(
            f"Invalid month: {month_name}"
        )

    if month_number >= 4:
        actual_year = start_year
    else:
        actual_year = start_year + 1

    month_text = datetime(
        actual_year,
        month_number,
        1
    ).strftime("%b")

    display_month = (
        f"{month_text}-{str(actual_year)[-2:]}"
    )

    return {
        "financial_year": financial_year,
        "month_name": month_name,
        "month_number": month_number,
        "year": actual_year,
        "display": display_month,
        "sort_key": (
            actual_year,
            month_number
        )
    }


# ============================================================
# VALIDATE WORKBOOK
# ============================================================

def validate_workbook(workbook):

    actual_sheets = {
        normalize_text(sheet).lower()
        for sheet in workbook.sheetnames
    }

    for required in REQUIRED_SHEETS:

        if normalize_text(required).lower() not in actual_sheets:

            raise ValueError(
                "Invalid file"
            )


# ============================================================
# CHECK DATA ROW
# ============================================================

HEADER_WORDS = {
    "invoice number",
    "invoice date",
    "invoice type",
    "invoice value(₹)",
    "invoice value",
    "note number",
    "note type",
    "note date",
    "note supply type",
    "note value (₹)",
    "gstin of supplier",
    "trade/legal name",
    "place of supply",
    "supply attract reverse charge",
    "taxable value (₹)",
    "taxable value",
    "integrated tax(₹)",
    "central tax(₹)",
    "state/ut tax(₹)",
    "cess(₹)",
    "tax amount",
    "remarks",
    "gstr-1/iff/gstr-5 period",
    "gstr-1/iff/gstr-5 filing date",
    "itc availability",
    "reason",
    "applicable % of tax rate"
}


def is_header_row(values):

    header_count = 0

    for value in values:

        text = normalize_text(
            value
        ).lower()

        if text in HEADER_WORDS:
            header_count += 1

    return header_count >= 2


def is_empty_row(values):

    return not any(
        value not in (None, "")
        for value in values
    )


# ============================================================
# GET DATA ROWS
# ============================================================

def get_data_rows(ws):

    rows = []

    for row_number in range(
        7,
        ws.max_row + 1
    ):

        values = [
            ws.cell(
                row_number,
                column
            ).value
            for column in range(
                1,
                ws.max_column + 1
            )
        ]

        if is_empty_row(values):
            continue

        if is_header_row(values):
            continue

        rows.append(values)

    return rows


# ============================================================
# B2B
# ============================================================

def process_b2b(ws, month):

    rows = []

    data_rows = get_data_rows(ws)

    for source in data_rows:

        source = source + [None] * (
            22 - len(source)
        )

        # 2024-25 files have Rate column at I,
        # so taxable value starts at J.
        #
        # 2025-26 files do not have Rate column,
        # so taxable value starts at I.

        if normalize_text(
            ws.cell(5, 9).value
        ).lower() == "rate(%)":

            gstin = source[0]
            trade = source[1]
            invoice_no = source[2]
            invoice_type = source[3]
            invoice_date = source[4]
            invoice_value = source[5]
            place = source[6]
            reverse = source[7]

            taxable = source[9]
            igst = source[10]
            cgst = source[11]
            sgst = source[12]
            cess = source[13]

            period = source[14]
            filing = source[15]
            itc = source[16]
            reason = source[17]
            rate = source[18]
            source_value = source[19]
            irn = source[20]
            irn_date = source[21]

        else:

            gstin = source[0]
            trade = source[1]
            invoice_no = source[2]
            invoice_type = source[3]
            invoice_date = source[4]
            invoice_value = source[5]
            place = source[6]
            reverse = source[7]

            taxable = source[8]
            igst = source[9]
            cgst = source[10]
            sgst = source[11]
            cess = source[12]

            period = source[13]
            filing = source[14]
            itc = source[15]
            reason = source[16]
            rate = source[17]
            source_value = source[18]
            irn = source[19]
            irn_date = source[20]

        total = (
            number(taxable)
            + number(igst)
            + number(cgst)
            + number(sgst)
            + number(cess)
        )

        rows.append([
            month,
            gstin,
            trade,
            invoice_no,
            invoice_type,
            invoice_date,
            invoice_value,
            place,
            reverse,
            taxable,
            igst,
            cgst,
            sgst,
            cess,
            total,
            period,
            filing,
            itc,
            reason,
            rate,
            source_value,
            irn,
            irn_date
        ])

    return rows


# ============================================================
# B2BA
# ============================================================

def process_b2ba(ws, month):

    rows = []

    data_rows = get_data_rows(ws)

    has_reduction_section = (
        ws.max_column >= 26
    )

    for source in data_rows:

        source = source + [None] * (
            26 - len(source)
        )

        # If row is just the secondary header,
        # don't process it.
        if is_header_row(source):
            continue

        original_invoice_no = source[0]
        original_invoice_date = source[1]
        gstin = source[2]
        trade = source[3]

        revised_invoice_no = source[4]
        revised_invoice_type = source[5]
        revised_invoice_date = source[6]
        invoice_value = source[7]

        place = source[8]
        reverse = source[9]

        if has_reduction_section:

            taxable = source[10]
            igst = source[11]
            cgst = source[12]
            sgst = source[13]
            cess = source[14]

            reduced = source[15]

            red_igst = source[16]
            red_cgst = source[17]
            red_sgst = source[18]
            red_cess = source[19]

            remarks = source[20]

            period = source[21]
            filing = source[22]
            itc = source[23]
            reason = source[24]
            rate = source[25]

        else:

            taxable = source[11]
            igst = source[12]
            cgst = source[13]
            sgst = source[14]
            cess = source[15]

            reduced = None
            red_igst = None
            red_cgst = None
            red_sgst = None
            red_cess = None
            remarks = None

            period = source[16]
            filing = source[17]
            itc = source[18]
            reason = source[19]
            rate = source[20]

        total = (
            number(taxable)
            + number(igst)
            + number(cgst)
            + number(sgst)
            + number(cess)
        )

        rows.append([
            month,
            original_invoice_no,
            original_invoice_date,
            gstin,
            trade,
            revised_invoice_no,
            revised_invoice_type,
            revised_invoice_date,
            invoice_value,
            place,
            reverse,
            taxable,
            igst,
            cgst,
            sgst,
            cess,
            total,
            reduced,
            red_igst,
            red_cgst,
            red_sgst,
            red_cess,
            remarks,
            period,
            filing,
            itc,
            reason,
            rate
        ])

    return rows


# ============================================================
# B2B-CDNR
# ============================================================

def process_b2b_cd_nr(ws, month):

    rows = []

    data_rows = get_data_rows(ws)

    has_reduction_section = (
        ws.max_column >= 28
    )

    for source in data_rows:

        source = source + [None] * (
            28 - len(source)
        )

        gstin = source[0]
        trade = source[1]

        note_number = source[2]
        note_type = source[3]
        note_supply_type = source[4]
        note_date = source[5]
        note_value = source[6]

        place = source[7]
        reverse = source[8]

        if has_reduction_section:

            taxable = source[9]

            igst = source[10]
            cgst = source[11]
            sgst = source[12]
            cess = source[13]

            reduced = source[14]

            red_igst = source[15]
            red_cgst = source[16]
            red_sgst = source[17]
            red_cess = source[18]

            remarks = source[19]

            period = source[20]
            filing = source[21]
            itc = source[22]
            reason = source[23]
            rate = source[24]
            source_value = source[25]
            irn = source[26]
            irn_date = source[27]

        else:

            taxable = source[10]

            igst = source[11]
            cgst = source[12]
            sgst = source[13]
            cess = source[14]

            reduced = None
            red_igst = None
            red_cgst = None
            red_sgst = None
            red_cess = None
            remarks = None

            period = source[15]
            filing = source[16]
            itc = source[17]
            reason = source[18]
            rate = source[19]
            source_value = source[20]
            irn = source[21]
            irn_date = source[22]

        total = (
            number(taxable)
            + number(igst)
            + number(cgst)
            + number(sgst)
            + number(cess)
        )

        rows.append([
            month,
            gstin,
            trade,
            note_number,
            note_type,
            note_supply_type,
            note_date,
            note_value,
            place,
            reverse,
            taxable,
            igst,
            cgst,
            sgst,
            cess,
            total,
            reduced,
            red_igst,
            red_cgst,
            red_sgst,
            red_cess,
            remarks,
            period,
            filing,
            itc,
            reason,
            rate,
            source_value,
            irn,
            irn_date
        ])

    return rows


# ============================================================
# B2B-CDNRA
# ============================================================

def process_b2b_cd_nra(ws, month):

    rows = []

    data_rows = get_data_rows(ws)

    has_reduction_section = (
        ws.max_column >= 28
    )

    for source in data_rows:

        source = source + [None] * (
            28 - len(source)
        )

        # Original details
        original_note_type = source[0]
        original_note_number = source[1]
        original_note_date = source[2]

        # Revised details
        gstin = source[3]
        trade = source[4]

        revised_note_number = source[5]
        revised_note_type = source[6]
        revised_supply_type = source[7]
        revised_note_date = source[8]
        note_value = source[9]

        place = source[10]
        reverse = source[11]

        if has_reduction_section:

            taxable = source[12]

            igst = source[13]
            cgst = source[14]
            sgst = source[15]
            cess = source[16]

            reduced = source[17]

            red_igst = source[18]
            red_cgst = source[19]
            red_sgst = source[20]
            red_cess = source[21]

            remarks = source[22]

            period = source[23]
            filing = source[24]
            itc = source[25]
            reason = source[26]
            rate = source[27]

        else:

            taxable = source[13]

            igst = source[14]
            cgst = source[15]
            sgst = source[16]
            cess = source[17]

            reduced = None
            red_igst = None
            red_cgst = None
            red_sgst = None
            red_cess = None
            remarks = None

            period = source[18]
            filing = source[19]
            itc = source[20]
            reason = source[21]
            rate = source[22]

        total = (
            number(taxable)
            + number(igst)
            + number(cgst)
            + number(sgst)
            + number(cess)
        )

        rows.append([
            month,

            original_note_type,
            original_note_number,
            original_note_date,

            gstin,
            trade,

            revised_note_number,
            revised_note_type,
            revised_supply_type,
            revised_note_date,
            note_value,

            place,
            reverse,

            taxable,
            igst,
            cgst,
            sgst,
            cess,
            total,

            reduced,
            red_igst,
            red_cgst,
            red_sgst,
            red_cess,

            remarks,
            period,
            filing,
            itc,
            reason,
            rate
        ])

    return rows


# ============================================================
# STYLE OUTPUT WORKBOOK
# ============================================================

def setup_sheet(ws, headers):

    # Header
    for column_number, header in enumerate(
        headers,
        start=1
    ):

        cell = ws.cell(
            row=1,
            column=column_number,
            value=header
        )

        cell.font = Font(
            name="Arial",
            size=12,
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="17365D"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )


    # Body default font
    for row in ws.iter_rows(
        min_row=2
    ):

        for cell in row:

            cell.font = Font(
                name="Arial",
                size=12,
                color="000000"
            )

            cell.alignment = Alignment(
                vertical="center"
            )


    # Freeze row 1
    ws.freeze_panes = "A2"


    # Filter
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}1"
    )


    # Row height
    ws.row_dimensions[1].height = 35


    # Column widths
    for column_number, header in enumerate(
        headers,
        start=1
    ):

        width = max(
            len(str(header)) + 3,
            15
        )

        width = min(
            width,
            35
        )

        ws.column_dimensions[
            get_column_letter(column_number)
        ].width = width


# ============================================================
# CREATE OUTPUT WORKBOOK
# ============================================================

def create_output_workbook():

    workbook = Workbook()

    default_sheet = workbook.active

    workbook.remove(
        default_sheet
    )


    sheets = [
        ("B2B", B2B_HEADERS),
        ("B2BA", B2BA_HEADERS),
        ("B2B-CDNR", B2B_CDNR_HEADERS),
        ("B2B-CDNRA", B2B_CDNRA_HEADERS)
    ]


    for sheet_name, headers in sheets:

        ws = workbook.create_sheet(
            sheet_name
        )

        setup_sheet(
            ws,
            headers
        )


    return workbook


# ============================================================
# APPEND DATA
# ============================================================

def append_rows(
    workbook,
    sheet_name,
    rows
):

    ws = workbook[sheet_name]

    for row in rows:

        ws.append(row)

        # Make sure appended cells use Arial 12
        for cell in ws[
            ws.max_row
        ]:

            cell.font = Font(
                name="Arial",
                size=12,
                color="000000"
            )


# ============================================================
# MERGE WORKER
# ============================================================

def merge_worker(
    task_id,
    uploaded_files
):

    task_dir = None

    try:

        with TASK_LOCK:

            TASKS[task_id]["status"] = "processing"
            TASKS[task_id]["progress"] = 0
            TASKS[task_id]["message"] = (
                "Reading files..."
            )


        # ----------------------------------------------------
        # Read all files first
        # ----------------------------------------------------

        prepared_files = []


        for file_path in uploaded_files:

            workbook = load_workbook(
                file_path,
                data_only=True,
                read_only=False
            )

            validate_workbook(
                workbook
            )

            period = get_period_from_readme(
                workbook
            )

            prepared_files.append({
                "path": file_path,
                "period": period
            })

            workbook.close()


        # ----------------------------------------------------
        # Sort chronologically
        # ----------------------------------------------------

        prepared_files.sort(
            key=lambda item:
                item["period"]["sort_key"]
        )


        total_files = len(
            prepared_files
        )


        # ----------------------------------------------------
        # Create output workbook
        # ----------------------------------------------------

        output_workbook = (
            create_output_workbook()
        )


        # ----------------------------------------------------
        # Process every file
        # ----------------------------------------------------

        for index, item in enumerate(
            prepared_files,
            start=1
        ):

            file_path = item["path"]
            period = item["period"]

            with TASK_LOCK:

                TASKS[task_id]["message"] = (
                    f"Processing file {index} "
                    f"of {total_files}: "
                    f"{period['display']}"
                )

                TASKS[task_id]["current_file"] = (
                    index
                )


            workbook = load_workbook(
                file_path,
                data_only=True,
                read_only=False
            )


            # ------------------------------------------------
            # B2B
            # ------------------------------------------------

            if "B2B" in workbook.sheetnames:

                rows = process_b2b(
                    workbook["B2B"],
                    period["display"]
                )

                append_rows(
                    output_workbook,
                    "B2B",
                    rows
                )


            # ------------------------------------------------
            # B2BA
            # ------------------------------------------------

            if "B2BA" in workbook.sheetnames:

                rows = process_b2ba(
                    workbook["B2BA"],
                    period["display"]
                )

                append_rows(
                    output_workbook,
                    "B2BA",
                    rows
                )


            # ------------------------------------------------
            # B2B-CDNR
            # ------------------------------------------------

            if "B2B-CDNR" in workbook.sheetnames:

                rows = process_b2b_cd_nr(
                    workbook["B2B-CDNR"],
                    period["display"]
                )

                append_rows(
                    output_workbook,
                    "B2B-CDNR",
                    rows
                )


            # ------------------------------------------------
            # B2B-CDNRA
            # ------------------------------------------------

            if "B2B-CDNRA" in workbook.sheetnames:

                rows = process_b2b_cd_nra(
                    workbook["B2B-CDNRA"],
                    period["display"]
                )

                append_rows(
                    output_workbook,
                    "B2B-CDNRA",
                    rows
                )


            workbook.close()


            # ------------------------------------------------
            # Progress
            # ------------------------------------------------

            progress = int(
                (index / total_files) * 100
            )

            with TASK_LOCK:

                TASKS[task_id]["progress"] = (
                    progress
                )

                TASKS[task_id]["message"] = (
                    f"{period['display']} completed"
                )


        # ----------------------------------------------------
        # Final styling
        # ----------------------------------------------------

        for ws in output_workbook.worksheets:

            # Reapply body formatting
            for row in ws.iter_rows(
                min_row=2
            ):

                for cell in row:

                    cell.font = Font(
                        name="Arial",
                        size=12,
                        color="000000"
                    )

            ws.freeze_panes = "A2"


        # ----------------------------------------------------
        # Save output
        # ----------------------------------------------------

        output_workbook.save(
            OUTPUT_FILE
        )

        output_workbook.close()


        # ----------------------------------------------------
        # Complete
        # ----------------------------------------------------

        with TASK_LOCK:

            TASKS[task_id]["status"] = "completed"

            TASKS[task_id]["progress"] = 100

            TASKS[task_id]["message"] = (
                "Files merged successfully."
            )

            TASKS[task_id]["output_file"] = (
                OUTPUT_FILE
            )


    except Exception as error:

        print(
            "GSTR-2B MERGE ERROR:",
            repr(error)
        )

        with TASK_LOCK:

            TASKS[task_id]["status"] = "error"

            TASKS[task_id]["message"] = (
                str(error)
            )

    finally:

        # ----------------------------------------------------
        # Delete temporary uploaded files
        # ----------------------------------------------------

        for file_path in uploaded_files:

            try:

                if os.path.exists(file_path):

                    os.remove(
                        file_path
                    )

            except Exception:
                pass


# ============================================================
# MAIN PAGE
# ============================================================

@gstr2bmerger_bp.route(
    "/tools/gstr2bmerger",
    methods=["GET"]
)
def gstr2bmerger():

    return render_template(
        "tools/gstr2bmerger.html"
    )


# ============================================================
# START MERGE
# ============================================================

@gstr2bmerger_bp.route(
    "/tools/gstr2bmerger",
    methods=["POST"]
)
def start_merge():

    files = request.files.getlist(
        "gstr2bFiles"
    )


    if not files:

        return jsonify({
            "success": False,
            "message": "No files uploaded."
        }), 400


    task_id = uuid.uuid4().hex


    temp_dir = tempfile.mkdtemp(
        prefix="gstr2b_"
    )


    uploaded_files = []


    try:

        for file in files:

            if not file or not file.filename:

                continue


            filename = secure_filename(
                file.filename
            )


            if not filename.lower().endswith(
                (".xlsx", ".xls")
            ):

                shutil.rmtree(
                    temp_dir,
                    ignore_errors=True
                )

                return jsonify({
                    "success": False,
                    "message": "Invalid file."
                }), 400


            file_path = os.path.join(
                temp_dir,
                filename
            )


            file.save(
                file_path
            )


            # Server-side validation
            workbook = load_workbook(
                file_path,
                read_only=True,
                data_only=True
            )

            validate_workbook(
                workbook
            )

            workbook.close()


            uploaded_files.append(
                file_path
            )


        if not uploaded_files:

            shutil.rmtree(
                temp_dir,
                ignore_errors=True
            )

            return jsonify({
                "success": False,
                "message": "No valid files uploaded."
            }), 400


        with TASK_LOCK:

            TASKS[task_id] = {

                "status": "starting",

                "progress": 0,

                "message": "Starting merge...",

                "current_file": 0,

                "total_files": len(
                    uploaded_files
                ),

                "output_file": None
            }


        thread = threading.Thread(
            target=merge_worker,
            args=(
                task_id,
                uploaded_files
            ),
            daemon=True
        )

        thread.start()


        return jsonify({

            "success": True,

            "task_id": task_id

        })


    except Exception as error:

        shutil.rmtree(
            temp_dir,
            ignore_errors=True
        )

        return jsonify({

            "success": False,

            "message": str(error)

        }), 400


# ============================================================
# PROGRESS
# ============================================================

@gstr2bmerger_bp.route(
    "/tools/gstr2bmerger/progress/<task_id>",
    methods=["GET"]
)
def merge_progress(task_id):

    with TASK_LOCK:

        task = TASKS.get(
            task_id
        )

        if not task:

            return jsonify({
                "success": False,
                "message": "Task not found."
            }), 404


        return jsonify({

            "success": True,

            "status": task["status"],

            "progress": task["progress"],

            "message": task["message"],

            "current_file": task[
                "current_file"
            ],

            "total_files": task[
                "total_files"
            ]

        })


# ============================================================
# DOWNLOAD PAGE
# ============================================================

@gstr2bmerger_bp.route(
    "/tools/gstr2bmerger/download-page/<task_id>",
    methods=["GET"]
)
def download_page(task_id):

    with TASK_LOCK:

        task = TASKS.get(
            task_id
        )


    if not task:

        return "Task not found.", 404


    if task["status"] != "completed":

        return "File is still being processed.", 400


    return render_template(
        "tools/gstr2bmergedownload.html",
        task_id=task_id
    )


# ============================================================
# DOWNLOAD FILE
# ============================================================

@gstr2bmerger_bp.route(
    "/tools/gstr2bmerger/download/<task_id>",
    methods=["GET"]
)
def download_merged_file(task_id):

    with TASK_LOCK:

        task = TASKS.get(
            task_id
        )


    if not task:

        return "Task not found.", 404


    output_file = task.get(
        "output_file"
    )


    if not output_file:

        return "Merged file not found.", 404


    if not os.path.exists(
        output_file
    ):

        return "Merged file not found.", 404


    return send_file(

        output_file,

        as_attachment=True,

        download_name=(
            "gstr2bmerged_taxdecipher.xlsx"
        ),

        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )